"""
Работа с Excel файлами для кураторов
"""
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config.settings import DATA_PATH, CURATORS


def get_curator_excel_path(curator: str) -> Path:
    """Получает путь к файлу Excel куратора"""
    curator_path = Path(DATA_PATH) / curator
    return curator_path / f"{curator}_participants.xlsx"


def create_or_update_curator_excel(
    curator: str,
    fio: str,
    inn: str,
    phone: str,
    curator_number: int,
    total_number: int,
    user_folder_path: str,
    pharmacy_name: str = "",
    pharmacy_number: str = "",
    position: str = ""
) -> bool:
    """
    Создает или обновляет Excel файл куратора с новым участником
    """
    try:
        excel_path = get_curator_excel_path(curator)
        
        # Проверяем, существует ли файл
        if excel_path.exists():
            wb = load_workbook(excel_path)
            ws = wb.active
            # Находим последнюю заполненную строку
            last_row = ws.max_row
            next_row = last_row + 1
        else:
            # Создаем новый файл
            wb = Workbook()
            ws = wb.active
            ws.title = curator
            
            # Создаем заголовки (без колонны "Путь")
            headers = ["№", "Общий №", "ФИО", "Аптека", "Номер аптеки", "Должность", "ИНН", "Телефон", "Дата регистрации"]
            
            # Стили для заголовков
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Устанавливаем ширину колонок
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 20
            
            next_row = 2
        
        # Добавляем новую строку с данными
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_data = [
            curator_number,                          # №
            total_number,                            # Общий №
            fio,                                     # ФИО
            pharmacy_name or "",                     # Аптека
            pharmacy_number or "",                   # Номер аптеки
            position or "",                          # Должность
            inn,                                     # ИНН
            phone,                                   # Телефон
            datetime.now().strftime("%d.%m.%Y %H:%M"),  # Дата регистрации
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_num)
            cell.value = value
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Выравнивание для чисел
            if col_num in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Сохраняем файл
        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"Ошибка при работе с Excel: {e}")
        return False


def get_all_curators_excel_stats() -> dict:
    """Получает статистику по всем куратором из их Excel файлов"""
    stats = {}
    
    for curator in CURATORS:
        excel_path = get_curator_excel_path(curator)
        if excel_path.exists():
            try:
                wb = load_workbook(excel_path)
                ws = wb.active
                # Количество строк минус заголовок
                count = ws.max_row - 1
                stats[curator] = count
            except:
                stats[curator] = 0
        else:
            stats[curator] = 0
    
    return stats


def create_or_update_general_excel(
    fio: str,
    inn: str,
    phone: str,
    curator: str,
    total_number: int,
    curator_number: int,
    pharmacy_name: str = "",
    pharmacy_number: str = "",
    position: str = ""
) -> bool:
    """
    Создает или обновляет общий Excel файл со всеми участниками
    """
    try:
        general_excel_path = Path(DATA_PATH) / "Все_участники.xlsx"
        
        # Проверяем, существует ли файл
        if general_excel_path.exists():
            wb = load_workbook(general_excel_path)
            ws = wb.active
            # Находим последнюю заполненную строку
            last_row = ws.max_row
            next_row = last_row + 1
        else:
            # Создаем новый файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Участники"
            
            # Создаем заголовки
            headers = ["Общий №", "№ у куратора", "ФИО", "Аптека", "Номер аптеки", "Должность", "ИНН", "Телефон", "Куратор", "Дата регистрации"]
            
            # Стили для заголовков
            header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Устанавливаем ширину колонок
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 20
            
            next_row = 2
        
        # Добавляем новую строку с данными
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_data = [
            total_number,                            # Общий №
            curator_number,                          # № у куратора
            fio,                                     # ФИО
            pharmacy_name or "",                     # Аптека
            pharmacy_number or "",                   # Номер аптеки
            position or "",                          # Должность
            inn,                                     # ИНН
            phone,                                   # Телефон
            curator,                                 # Куратор
            datetime.now().strftime("%d.%m.%Y %H:%M"),  # Дата регистрации
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_num)
            cell.value = value
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Выравнивание для чисел
            if col_num in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Сохраняем файл
        wb.save(general_excel_path)
        return True
    except Exception as e:
        print(f"Ошибка при работе с общим Excel: {e}")
        return False
